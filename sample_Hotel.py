#20190401 
# https://note.nkmk.me/python-openpyxl-usage/
'''
Auto form fulling test_Hotel_reservation_form using selsenium.


'''
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl 

class EX_LIST:
    def __init__(self,file_1):
        self.file_1 = file_1
    def get_value_list(self,t_2d):
        return([[cell.value for cell in row] for row in t_2d])


    def get_list_2d(self,sheet, start_row, end_row, start_col, end_col):
        return self.get_value_list(sheet.iter_rows(min_row=start_row,
                                          max_row=end_row,
                                          min_col=start_col,
                                          max_col=end_col))
    
    def make_list(self):
        wb = openpyxl.load_workbook(self.file_1)
        sheet_names = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(sheet_names[0])     
#        print(sheet.max_row)
        l_2d = self.get_list_2d(sheet, 2, sheet.max_row, 1, sheet.max_column)
        return l_2d
class FORM_SEND():
    
    def __init__(self, driver):
        self._driver = driver
    
    def name_formsend_befor_clear(self,name,text1):
        try:
            element =driver.find_element_by_name(name)
            element.clear()
            self._driver.find_element_by_name(name).send_keys(text1)
#            elememt.send_keys(text1)
        except KeyError: # KerError 
            """
            except Exception as e:
                print("例外args:", e.args)
            """
            time.sleep(3)
            self._driver.quit()
            
    def determin_botun(self,text4):
        try:
            self._driver.find_element_by_id(text4).click()
        except  Exception as e:
            print("Exception args:", e.args) 
            
if __name__=='__main__':
    test1 = EX_LIST('selenium_test.xlsx')
    result =test1.make_list()
    print(result)
    url_1= 'http://example.selenium.jp/reserveApp/'
    driver = webdriver.Chrome()
    driver.get(url_1)

    form_ex1 = FORM_SEND(driver)
    for row in range(len(result)):
        # redserve date 
        form_ex1.name_formsend_befor_clear('reserve_y',result[row][0])
        time.sleep(1)
        # manth
        form_ex1.name_formsend_befor_clear('reserve_m',result[row][1])
        time.sleep(1)
        # day
        form_ex1.name_formsend_befor_clear('reserve_d',result[row][2])
        time.sleep(1)
        # How long stay 
        form_ex1.name_formsend_befor_clear('reserve_t',result[row][3])
        time.sleep(1)
        # How manay person
        form_ex1.name_formsend_befor_clear('hc',result[row][4])
        time.sleep(1)
        # have a breakfasr 
        if result[row][6] =='1':
            form_ex1.determin_botun('breakfast_on')
        else:
            form_ex1.determin_botun('breakfast_off')
        #  plan a or b
        if result[row][8] =='1':
            form_ex1.determin_botun('plan_a')
        else:
            form_ex1.determin_botun('plan_b')
        # customer name    
        form_ex1.name_formsend_befor_clear('gname',result[row][5])
        time.sleep(1)
        # next page
        form_ex1.determin_botun('goto_next')
        time.sleep(3)
        form_ex1.determin_botun('commit')
        time.sleep(3)
        form_ex1.determin_botun('returnto_checkInfo')
        form_ex1.determin_botun('returnto_index')
        time.sleep(3)
    time.sleep(2)
    driver.quit()    
